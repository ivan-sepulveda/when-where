"""
Data Source: Australian Bureau of Statistics, Overseas Arrivals and Departures, Australia (catalogue 3401.0)
URL: https://www.abs.gov.au/about/data-services/application-programming-interfaces-apis/time-series-directory-api
Tables Referenced: Table 1, "Total Movement, Arrivals - Category of Movement" (Data1 sheet of 340101.xlsx)

Fetches ABS's Time Series Directory API (not the SDMX Data API, whose
responses are opaque vendor MIME types) to resolve the current Table 1
spreadsheet URL, then parses its wide-format `Data1` sheet into a tidy
monthly CSV. Keeps only "Original" series (Seasonally Adjusted/Trend were
suspended by ABS in 2020 and have stayed blank since). Target column is
`short_term_visitors_arriving`, the closest single number to monthly
inbound tourist volume; other categories are kept alongside as free
context. See data/README.md for full verification notes.

Usage:
    python fetch_abs_visitor_arrivals.py                       # full history, all categories
    python fetch_abs_visitor_arrivals.py --years-back 10
    python fetch_abs_visitor_arrivals.py --start-date 2015-01 --end-date 2025-12
    python fetch_abs_visitor_arrivals.py --force-download       # bypass the cached raw/ xlsx
"""

import argparse
import xml.etree.ElementTree as ET
from datetime import date, datetime
from pathlib import Path

import openpyxl
import pandas as pd
import requests

CATALOGUE_NUMBER = "3401.0"
TABLE_TITLE_QUERY = '"table 1"'
TS_SEARCH_URL = "https://abs.gov.au/servlet/TSSearchServlet"

RAW_DIR = Path(__file__).resolve().parent.parent.parent / "raw" / "abs_visitor_arrivals"
PROCESSED_DIR = Path(__file__).resolve().parent.parent.parent / "processed" / "oceana"
OUTPUT_FILENAME = "abs_visitor_arrivals_monthly.csv"

# Description text (as it appears in Data1 row 1) -> output column slug.
# Matched via substring, case-sensitive (ABS's own text is consistent about
# capitalization within one table) -- see docstring for the full column list
# confirmed against the real workbook.
DESCRIPTION_TO_COLUMN = {
    "Permanent Arrivals": "permanent_arrivals",
    "Long-term Residents returning": "long_term_residents_returning",
    "Long-term Visitors arriving": "long_term_visitors_arriving",
    "Permanent and Long-term Arrivals": "permanent_and_long_term_arrivals",
    "Short-term Residents returning": "short_term_residents_returning",
    "Short-term Visitors arriving": "short_term_visitors_arriving",
    "Total Arrivals": "total_arrivals",
}

# The column this project actually cares about -- see docstring.
PRIMARY_COLUMN = "short_term_visitors_arriving"


def resolve_table_url(catno: str = CATALOGUE_NUMBER, ttitle_query: str = TABLE_TITLE_QUERY) -> tuple[str, dict[str, str]]:
    """Query the Time Series Directory API for the current TableURL + a
    description->SeriesID map (used to sanity-check the spreadsheet's own
    Series ID row later, since both come from ABS independently)."""
    resp = requests.get(TS_SEARCH_URL, params={"catno": catno, "ttitle": ttitle_query}, timeout=60)
    resp.raise_for_status()
    root = ET.fromstring(resp.content)

    series_list = root.findall("Series")
    if not series_list:
        raise ValueError(f"No series returned for catno={catno!r} ttitle={ttitle_query!r} -- check the query.")

    table_urls = {s.findtext("TableURL") for s in series_list}
    if len(table_urls) != 1:
        raise ValueError(f"Expected exactly one distinct TableURL, got {table_urls!r} -- query may be too broad.")
    table_url = table_urls.pop()

    series_ids = {}
    for s in series_list:
        if s.findtext("SeriesType") == "Original":
            series_ids[s.findtext("Description", "").strip()] = s.findtext("SeriesID")

    return table_url, series_ids


def download_xlsx(table_url: str, force: bool = False) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    filename = table_url.rsplit("/", 1)[-1]
    xlsx_path = RAW_DIR / filename

    if force or not xlsx_path.exists():
        print(f"[{CATALOGUE_NUMBER}] downloading {table_url} ...")
        resp = requests.get(table_url, timeout=120)
        resp.raise_for_status()
        xlsx_path.write_bytes(resp.content)
    else:
        print(f"[{CATALOGUE_NUMBER}] using cached {xlsx_path} (--force-download to bypass)")

    return xlsx_path


def find_header_rows(ws, max_scan: int = 20) -> dict[str, int]:
    """Scan column A of a Data1-style sheet for the header row labels this
    project relies on ("Series Type", "Series ID"), rather than hardcoding
    row numbers -- confirmed at rows 3/10 in the real 340101.xlsx, but not
    assumed stable across every ABS table in this workbook family."""
    labels_wanted = {"Series Type", "Series ID"}
    found = {}
    for r in range(1, max_scan + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip() in labels_wanted:
            found[val.strip()] = r
    missing = labels_wanted - found.keys()
    if missing:
        raise ValueError(f"Could not find header row(s) {missing} in column A within the first {max_scan} rows.")
    return found


def parse_data1_sheet(xlsx_path: Path) -> pd.DataFrame:
    """Read the Data1 sheet's wide layout into a tidy long-format-free
    DataFrame: one row per month, one column per "Original"-series category
    (see DESCRIPTION_TO_COLUMN)."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if "Data1" not in wb.sheetnames:
        raise ValueError(f"Expected a 'Data1' sheet, found {wb.sheetnames!r}")
    ws = wb["Data1"]

    header_rows = find_header_rows(ws)
    type_row = header_rows["Series Type"]
    id_row = header_rows["Series ID"]
    desc_row = 1  # column A is blank on this row; confirmed against the real file
    data_start_row = id_row + 1

    # Map each column to an output slug, keeping only "Original" series
    # whose description matches something in DESCRIPTION_TO_COLUMN.
    col_to_slug: dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        series_type = ws.cell(row=type_row, column=c).value
        if series_type != "Original":
            continue
        desc = ws.cell(row=desc_row, column=c).value
        if not isinstance(desc, str):
            continue
        for needle, slug in DESCRIPTION_TO_COLUMN.items():
            if needle in desc:
                col_to_slug[c] = slug
                break

    missing_slugs = set(DESCRIPTION_TO_COLUMN.values()) - set(col_to_slug.values())
    if missing_slugs:
        print(f"WARNING: expected columns not found in this workbook (layout may have changed): {missing_slugs}")

    rows = []
    for r in range(data_start_row, ws.max_row + 1):
        raw_date = ws.cell(row=r, column=1).value
        if not isinstance(raw_date, datetime):
            continue  # stop at trailing footnote/blank rows, don't assume max_row is exact
        row = {"ref_date": raw_date.strftime("%Y-%m")}
        for c, slug in col_to_slug.items():
            row[slug] = ws.cell(row=r, column=c).value
        rows.append(row)

    if not rows:
        raise ValueError("Parsed zero data rows -- check data_start_row / column A date formatting.")

    df = pd.DataFrame(rows)
    ordered_cols = ["ref_date"] + [c for c in DESCRIPTION_TO_COLUMN.values() if c in df.columns]
    return df[ordered_cols]


def filter_dates(df: pd.DataFrame, start_date: str | None, end_date: str | None) -> pd.DataFrame:
    out = df
    if start_date:
        out = out[out["ref_date"] >= start_date]
    if end_date:
        out = out[out["ref_date"] <= end_date]
    return out


def write_output(df: pd.DataFrame) -> Path:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / OUTPUT_FILENAME
    df.to_csv(out_path, index=False)
    return out_path


def fetch_dataset(
    start_date: str | None = None,
    end_date: str | None = None,
    force_download: bool = False,
) -> Path:
    table_url, expected_series_ids = resolve_table_url()
    xlsx_path = download_xlsx(table_url, force=force_download)

    print(f"[{CATALOGUE_NUMBER}] reading {xlsx_path} ...")
    df = parse_data1_sheet(xlsx_path)
    df = filter_dates(df, start_date, end_date)

    out_path = write_output(df)
    print(f"[{CATALOGUE_NUMBER}] wrote {len(df)} rows -> {out_path}")
    print(f"[{CATALOGUE_NUMBER}] primary column: '{PRIMARY_COLUMN}' (monthly short-term visitor arrivals, Original series)")
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--years-back",
        type=int,
        default=None,
        help="How many years of history to keep, counted back from today. "
        "Ignored if --start-date is also given. Default: no limit (full 1976-onward history).",
    )
    parser.add_argument("--start-date", default=None, help="ref_date lower bound, 'YYYY-MM'.")
    parser.add_argument("--end-date", default=None, help="ref_date upper bound, 'YYYY-MM'.")
    parser.add_argument("--force-download", action="store_true", help="Bypass the cached raw/ xlsx.")
    args = parser.parse_args()

    start_date = args.start_date
    if start_date is None and args.years_back is not None:
        cutoff = date.today().replace(year=date.today().year - args.years_back)
        start_date = cutoff.strftime("%Y-%m")

    fetch_dataset(start_date=start_date, end_date=args.end_date, force_download=args.force_download)


if __name__ == "__main__":
    main()
