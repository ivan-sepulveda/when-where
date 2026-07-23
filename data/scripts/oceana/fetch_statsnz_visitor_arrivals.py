"""
Data Source: Stats NZ, International travel release (monthly)
URL: https://www.stats.govt.nz/information-releases/international-travel-may-2026/
Tables Referenced: Table 1, "Monthly visitor arrivals" (sheet 'Tables 1&2')

Fetches Table 1 into a tidy monthly CSV, one workbook per release month.
The table layout is a small report block, not a plain grid: a fiscal-year
row (5 columns, e.g. "2021/22".."2025/26", each meaning "year ended 31
May") crossed with 12 month rows (Jun-May, NZ's tourism year).
`month_to_ref_date()` converts that fiscal-year x month grid into ordered
`YYYY-MM` dates, since the same column means a different calendar year
depending on the month row. YoY change columns apply only to the most
recent fiscal year. See data/README.md for full layout and verification
notes.

Usage:
    python fetch_statsnz_visitor_arrivals.py                         # default: May 2026 release
    python fetch_statsnz_visitor_arrivals.py --release-year 2026 --release-month 5
    python fetch_statsnz_visitor_arrivals.py --url "https://.../some-other-release.xlsx"
    python fetch_statsnz_visitor_arrivals.py --force-download        # bypass the cached raw/ xlsx
"""

import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd
import requests

RAW_DIR = Path(__file__).resolve().parent.parent.parent / "raw" / "statsnz_visitor_arrivals"
PROCESSED_DIR = Path(__file__).resolve().parent.parent.parent / "processed" / "oceana"
OUTPUT_FILENAME = "statsnz_visitor_arrivals_monthly.csv"

SHEET_NAME = "Tables 1&2"
TABLE_TITLE = "Table 1"
FOOTER_TEXT = "Source: Stats NZ"

# Month abbreviation (as it appears in col A) -> (calendar month number,
# which half of the "YYYY/YY" fiscal-year label it belongs to).
# NZ's tourism year runs Jun -> May: Jun-Dec fall in the fiscal label's
# first calendar year, Jan-May fall in its second.
MONTH_ORDER = [
    ("Jun", 6, "first"), ("Jul", 7, "first"), ("Aug", 8, "first"),
    ("Sep", 9, "first"), ("Oct", 10, "first"), ("Nov", 11, "first"),
    ("Dec", 12, "first"),
    ("Jan", 1, "second"), ("Feb", 2, "second"), ("Mar", 3, "second"),
    ("Apr", 4, "second"), ("May", 5, "second"),
]
MONTH_TO_INFO = {name: (num, half) for name, num, half in MONTH_ORDER}

FISCAL_YEAR_RE = re.compile(r"^(\d{4})/(\d{2})$")

DEFAULT_RELEASE_YEAR = 2026
DEFAULT_RELEASE_MONTH = 5  # May


def build_url_for_release(year: int, month: int) -> str:
    """Construct the Stats NZ download URL for a given release (year, month).
    Verified by pattern-matching against the real May-2026 URL the user
    supplied; not yet confirmed for other months/years since stats.govt.nz
    is unreachable from this sandbox (see module docstring)."""
    month_name = pd.Timestamp(year=year, month=month, day=1).strftime("%B")
    month_slug = month_name.lower()
    return (
        "https://www.stats.govt.nz/assets/Uploads/International-travel/"
        f"International-travel-{month_name}-{year}/Download-data/"
        f"international-visitor-arrivals-to-new-zealand-{month_slug}-{year}.xlsx"
    )


def download_xlsx(url: str, force: bool = False) -> Path:
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    filename = url.rsplit("/", 1)[-1]
    xlsx_path = RAW_DIR / filename

    if force or not xlsx_path.exists():
        print(f"[stats.govt.nz] downloading {url} ...")
        resp = requests.get(url, timeout=120)
        resp.raise_for_status()
        xlsx_path.write_bytes(resp.content)
    else:
        print(f"[stats.govt.nz] using cached {xlsx_path} (--force-download to bypass)")

    return xlsx_path


def find_table1_header_rows(ws, max_scan: int = 40) -> dict[str, int]:
    """Scan column A for 'Table 1' then, below it, 'Month' -- the header
    row that anchors the fiscal-year row and subheader row beneath it.
    Row numbers aren't hardcoded since Stats NZ has reshuffled rows
    (notes, blank spacers) across past releases."""
    table1_row = None
    for r in range(1, max_scan + 1):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip() == TABLE_TITLE:
            table1_row = r
            break
    if table1_row is None:
        raise ValueError(f"Could not find '{TABLE_TITLE}' in column A of sheet '{SHEET_NAME}'.")

    month_header_row = None
    for r in range(table1_row + 1, table1_row + 10):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip() == "Month":
            month_header_row = r
            break
    if month_header_row is None:
        raise ValueError(f"Could not find 'Month' header row below row {table1_row}.")

    fiscal_year_row = month_header_row + 1
    subheader_row = month_header_row + 2

    # Data doesn't necessarily start immediately after the subheader row --
    # the real file has a blank spacer row in between -- so scan forward
    # for the first row whose column A is a recognized month abbreviation
    # rather than assuming a fixed offset.
    data_start_row = None
    for r in range(subheader_row + 1, subheader_row + 6):
        val = ws.cell(row=r, column=1).value
        if isinstance(val, str) and val.strip() in MONTH_TO_INFO:
            data_start_row = r
            break
    if data_start_row is None:
        raise ValueError(f"Could not find first month data row below row {subheader_row}.")

    return {
        "table1_row": table1_row,
        "month_header_row": month_header_row,
        "fiscal_year_row": fiscal_year_row,
        "subheader_row": subheader_row,
        "data_start_row": data_start_row,
    }


def parse_table1(xlsx_path: Path) -> pd.DataFrame:
    """Read Table 1 (Monthly visitor arrivals) out of the 'Tables 1&2'
    sheet and return a tidy monthly DataFrame: one row per real calendar
    month, columns ref_date / fiscal_year / month / visitor_arrivals /
    change_number_yoy / change_percent_yoy."""
    wb = openpyxl.load_workbook(xlsx_path, data_only=True)
    if SHEET_NAME not in wb.sheetnames:
        raise ValueError(f"Expected a '{SHEET_NAME}' sheet, found {wb.sheetnames!r}")
    ws = wb[SHEET_NAME]

    rows_meta = find_table1_header_rows(ws)
    fy_row = rows_meta["fiscal_year_row"]
    subheader_row = rows_meta["subheader_row"]
    data_start_row = rows_meta["data_start_row"]

    # Locate the fiscal-year columns (e.g. "2021/22" .. "2025/26"), in
    # left-to-right / chronological order.
    fy_cols: dict[int, str] = {}
    for c in range(2, ws.max_column + 1):
        val = ws.cell(row=fy_row, column=c).value
        if isinstance(val, str) and FISCAL_YEAR_RE.match(val.strip()):
            fy_cols[c] = val.strip()
    if not fy_cols:
        raise ValueError(f"No fiscal-year columns (e.g. '2021/22') found in row {fy_row}.")

    most_recent_fy = fy_cols[max(fy_cols)]  # rightmost column = most recent

    # Locate the YoY change columns via the subheader row's "Number"/"Percent" labels.
    change_number_col = change_percent_col = None
    for c in range(max(fy_cols) + 1, ws.max_column + 1):
        val = ws.cell(row=subheader_row, column=c).value
        if val == "Number":
            change_number_col = c
        elif val == "Percent":
            change_percent_col = c
    if change_number_col is None or change_percent_col is None:
        raise ValueError(f"Could not find 'Number'/'Percent' change columns in row {subheader_row}.")

    records = []
    r = data_start_row
    while True:
        month_name = ws.cell(row=r, column=1).value
        if not isinstance(month_name, str) or month_name.strip() not in MONTH_TO_INFO:
            break  # hit the "Source: Stats NZ" footer (or ran out of expected months)
        month_name = month_name.strip()
        month_num, half = MONTH_TO_INFO[month_name]

        change_number = ws.cell(row=r, column=change_number_col).value
        change_percent = ws.cell(row=r, column=change_percent_col).value

        for c, fy_label in fy_cols.items():
            year_start, year_end_suffix = FISCAL_YEAR_RE.match(fy_label).groups()
            calendar_year = int(year_start) if half == "first" else int(year_start[:2] + year_end_suffix)
            ref_date = f"{calendar_year}-{month_num:02d}"

            is_most_recent_fy = fy_label == most_recent_fy
            records.append({
                "ref_date": ref_date,
                "fiscal_year": fy_label,
                "month": month_name,
                "visitor_arrivals": ws.cell(row=r, column=c).value,
                "change_number_yoy": change_number if is_most_recent_fy else None,
                "change_percent_yoy": change_percent if is_most_recent_fy else None,
            })
        r += 1

    if not records:
        raise ValueError("Parsed zero data rows -- check data_start_row / month labels in column A.")

    df = pd.DataFrame(records).sort_values("ref_date").reset_index(drop=True)
    return df


def filter_dates(df: pd.DataFrame, start_date: str | None, end_date: str | None) -> pd.DataFrame:
    out = df
    if start_date:
        out = out[out["ref_date"] >= start_date]
    if end_date:
        out = out[out["ref_date"] <= end_date]
    return out


def write_output(df: pd.DataFrame) -> Path:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / OUTPUT_FILENAME
    df.to_csv(out_path, index=False)
    return out_path


def fetch_dataset(
    url: str | None = None,
    release_year: int = DEFAULT_RELEASE_YEAR,
    release_month: int = DEFAULT_RELEASE_MONTH,
    start_date: str | None = None,
    end_date: str | None = None,
    force_download: bool = False,
) -> Path:
    resolved_url = url or build_url_for_release(release_year, release_month)
    xlsx_path = download_xlsx(resolved_url, force=force_download)

    print(f"[stats.govt.nz] reading {xlsx_path} ...")
    df = parse_table1(xlsx_path)
    df = filter_dates(df, start_date, end_date)

    out_path = write_output(df)
    print(f"[stats.govt.nz] wrote {len(df)} rows -> {out_path}")
    print("[stats.govt.nz] primary column: 'visitor_arrivals' (Table 1, Monthly visitor arrivals)")
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--url", default=None, help="Exact xlsx URL to fetch. Overrides --release-year/--release-month.")
    parser.add_argument("--release-year", type=int, default=DEFAULT_RELEASE_YEAR, help=f"Release year, default {DEFAULT_RELEASE_YEAR}.")
    parser.add_argument("--release-month", type=int, default=DEFAULT_RELEASE_MONTH, help=f"Release month (1-12), default {DEFAULT_RELEASE_MONTH} (May).")
    parser.add_argument("--start-date", default=None, help="ref_date lower bound, 'YYYY-MM'.")
    parser.add_argument("--end-date", default=None, help="ref_date upper bound, 'YYYY-MM'.")
    parser.add_argument("--force-download", action="store_true", help="Bypass the cached raw/ xlsx.")
    args = parser.parse_args()

    fetch_dataset(
        url=args.url,
        release_year=args.release_year,
        release_month=args.release_month,
        start_date=args.start_date,
        end_date=args.end_date,
        force_download=args.force_download,
    )


if __name__ == "__main__":
    main()
