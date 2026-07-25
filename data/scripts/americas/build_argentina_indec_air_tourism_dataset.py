"""
Data Source: INDEC (Instituto Nacional de Estadística y Censos de la
República Argentina), "Turismo receptivo y emisivo. Series original,
desestacionalizada y tendencia-ciclo. Vía aérea internacional"
Page: https://www.indec.gob.ar/indec/web/Nivel4-Tema-3-13-55
URL: https://www.indec.gob.ar/ftp/cuadros/economia/series_eti_via_aerea.xlsx

Fetches the workbook (Enero 2016-Mayo 2026 edition as of writing) and
parses it into a tidy long-format monthly CSV covering both directions
of international air travel: "Turismo receptivo" (foreign visitors
arriving in Argentina by air -- the destination-relevant series) and
"Turismo emisivo" (Argentine residents departing by air).
`download_workbook()` attempts a live fetch first (browser User-Agent,
since a plain default UA is a common 403 cause on Argentine/Chilean
government sites in this project -- same fix as Chile INE elsewhere in
this file) and falls back to the cached copy in raw/ if that fails (as
it does in this sandbox -- indec.gob.ar is network-blocked here,
confirmed, same issue as ine.gob.cl/StatCan/Eurostat/e-Stat).

Each sheet carries three parallel series per month: "Serie original" (raw
monthly count, in miles/thousands), "Serie desestacionalizada" (seasonally
adjusted), and "Tendencia-ciclo" (trend-cycle, smoothed). See
data/README.md for why "Serie original" -- not the seasonally adjusted
series -- is the one wired into compute_peak_tourism_indicator.py: a peak
season indicator needs the seasonal signal, not one with it removed.

Usage:
    python build_argentina_indec_air_tourism_dataset.py
    python build_argentina_indec_air_tourism_dataset.py --force-download
"""

import argparse
from pathlib import Path

import openpyxl
import pandas as pd
import requests

RAW_DIR = Path(__file__).resolve().parent.parent.parent / "raw" / "argentina_indec"
PROCESSED_DIR = Path(__file__).resolve().parent.parent.parent / "processed" / "americas"
CACHED_XLSX_PATH = RAW_DIR / "series_eti_via_aerea.xlsx"
OUTPUT_FILENAME = "argentina_indec_air_tourism_monthly.csv"

SOURCE_PAGE_URL = "https://www.indec.gob.ar/indec/web/Nivel4-Tema-3-13-55"
SOURCE_XLSX_URL = "https://www.indec.gob.ar/ftp/cuadros/economia/series_eti_via_aerea.xlsx"
ATTRIBUTION = f"INDEC, Dirección de Estadísticas Básicas de la Balanza de Pagos -- {SOURCE_PAGE_URL}"

# Sheet name -> flow label used in the output ("receptivo" = foreign
# visitors arriving in Argentina, "emisivo" = Argentine residents departing).
SHEETS = {
    "Turismo receptivo": "receptivo",
    "Turismo emisivo": "emisivo",
}

DATE_COL = 2       # column B: "Período" (first-of-month date)
ORIGINAL_COL = 3   # column C: Serie original (miles)
SEASONAL_COL = 4   # column D: Serie desestacionalizada (miles)
TREND_COL = 5      # column E: Tendencia-ciclo (miles)
DATA_START_ROW = 5


def _download_live() -> None:
    """GET SOURCE_XLSX_URL and overwrite CACHED_XLSX_PATH. A plain default
    User-Agent gets a 403 from indec.gob.ar (bot-blocking) -- a
    browser-like one works, same fix used for Chile INE/SimpleMaps
    elsewhere in this project."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        )
    }
    resp = requests.get(SOURCE_XLSX_URL, headers=headers, timeout=60)
    resp.raise_for_status()
    RAW_DIR.mkdir(parents=True, exist_ok=True)
    CACHED_XLSX_PATH.write_bytes(resp.content)


def download_workbook(force: bool = False) -> Path:
    """Fetch the ETI vía aérea workbook, preferring a live download but
    falling back to the cached copy in raw/ if the request fails (as it
    does in this sandbox -- indec.gob.ar is network-blocked here, see
    module docstring). With force=True there's no fallback -- a failed
    request raises, since the whole point of --force-download is to
    bypass the cache."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    if force:
        _download_live()
        print(f"Downloaded (forced) -> {CACHED_XLSX_PATH}")
        return CACHED_XLSX_PATH

    try:
        _download_live()
        print(f"Downloaded -> {CACHED_XLSX_PATH}")
        return CACHED_XLSX_PATH
    except requests.RequestException as e:
        if CACHED_XLSX_PATH.exists():
            print(f"Live download failed ({e}) -- using cached copy: {CACHED_XLSX_PATH}")
            return CACHED_XLSX_PATH
        raise


def load_workbook(path: Path) -> openpyxl.Workbook:
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


def parse_sheet(ws, flow: str) -> pd.DataFrame:
    """Parse one sheet into tidy rows. Stops at the first blank date cell
    (row 130 in the source: a spacer row before the "Fuente:" footnote)."""
    rows = []
    r = DATA_START_ROW
    while True:
        date_val = ws.cell(row=r, column=DATE_COL).value
        if date_val is None:
            break

        original = ws.cell(row=r, column=ORIGINAL_COL).value
        seasonal = ws.cell(row=r, column=SEASONAL_COL).value
        trend = ws.cell(row=r, column=TREND_COL).value

        rows.append({
            "flow": flow,
            "ref_date": f"{date_val.year:04d}-{date_val.month:02d}",
            "original_thousands": original,
            "seasonally_adjusted_thousands": seasonal,
            "trend_cycle_thousands": trend,
            # Whole-passenger count from the raw (non-seasonally-adjusted)
            # series -- the "Miles" unit is thousands, so *1000. Rounded
            # since the source only carries ~6 significant figures.
            "passengers": round(original * 1000) if isinstance(original, (int, float)) else None,
        })
        r += 1

    return pd.DataFrame(rows)


def build_dataset(xlsx_path: Path) -> pd.DataFrame:
    wb = load_workbook(xlsx_path)
    frames = []
    for sheet_name, flow in SHEETS.items():
        if sheet_name not in wb.sheetnames:
            raise ValueError(f"{xlsx_path} has no sheet {sheet_name!r} -- is this really the ETI vía aérea workbook?")
        frames.append(parse_sheet(wb[sheet_name], flow))

    df = pd.concat(frames, ignore_index=True)
    return df.sort_values(["flow", "ref_date"]).reset_index(drop=True)


def write_output(df: pd.DataFrame) -> Path:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    out_path = PROCESSED_DIR / OUTPUT_FILENAME
    df.to_csv(out_path, index=False)
    return out_path


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument("--force-download", action="store_true", help="Bypass the cached raw/ copy.")
    args = parser.parse_args()

    xlsx_path = download_workbook(force=args.force_download)
    print(f"Reading {xlsx_path} ...")
    df = build_dataset(xlsx_path)
    out_path = write_output(df)
    print(f"Wrote {len(df)} rows -> {out_path}")

    for flow in df["flow"].unique():
        sub = df[df["flow"] == flow]
        print(f"  {flow}: {sub['ref_date'].min()} - {sub['ref_date'].max()} ({len(sub)} months)")

    receptivo = df[df["flow"] == "receptivo"]
    peak = receptivo.loc[receptivo["passengers"].idxmax()]
    print(f"Sanity check -- receptivo peak month: {peak['ref_date']} at {peak['passengers']:,} passengers")


if __name__ == "__main__":
    main()
