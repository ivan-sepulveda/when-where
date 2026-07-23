"""
Data Source: Chile INE (Instituto Nacional de Estadísticas), Encuesta Mensual de Alojamiento Turística (EMAT)
URL: https://www.ine.gob.cl/estadisticas-por-tema/comercio-y-servicios/actividad-mensual-del-turismo
Tables Referenced: Workbook sheets "1"-"33" (33 monthly time-series
tables, one metric each -- e.g. Cuadro 1 = overnight stays, Cuadro 6 =
arrivals); sheet "34" is a static region/destino-turístico/comuna lookup,
not a time series.

Fetches the EMAT workbook and writes a tidy long-format CSV of one or
more tables, plus the region/comuna reference table. Defaults to Table 1
(overnight stays) -- see data/README.md for why that's the recommended
indicator over arrivals. The region/destino turístico hierarchy is
conveyed only through bold formatting, not indentation, and end-of-table
detection ignores literal "-" placeholder cells for destinos added
mid-series -- both handled internally, see data/README.md for the full
parsing rationale. Live fetching is blocked in this sandbox, so the
cached raw copy is used as a fallback.

Usage:
    python fetch_chile_ine_tourism_accommodation.py                  # Cuadro 1 only (overnight stays, total)
    python fetch_chile_ine_tourism_accommodation.py --table 6        # Cuadro 6 (arrivals, total)
    python fetch_chile_ine_tourism_accommodation.py --all-tables     # every Cuadro 1-33 in one long CSV
    python fetch_chile_ine_tourism_accommodation.py --list-tables    # print all 34 table titles and exit
    python fetch_chile_ine_tourism_accommodation.py --force-download # bypass the cached raw/ copy
"""

import argparse
import re
from pathlib import Path

import openpyxl
import pandas as pd
import requests

# ---------------------------------------------------------------------------
# Config -- the only section you should need to edit.
# ---------------------------------------------------------------------------

SOURCE_PAGE_URL = "https://www.ine.gob.cl/estadisticas-por-tema/comercio-y-servicios/actividad-mensual-del-turismo"
SOURCE_XLSX_URL = (
    "https://www.ine.gob.cl/docs/default-source/actividad-del-turismo/"
    "cuadros-estadisticos-dos/serie-hist%C3%B3rica-metodolog%C3%ADa-2017/"
    "series-mensuales-de-julio-2016-a-la-fecha.xlsx"
)

# Cuadro number to parse by default when neither --table nor --all-tables is
# given. 1 = "Número de pernoctaciones de pasajeros en establecimientos de
# alojamiento turístico" (overnight stays, total) -- see docstring for why
# overnight stays is the recommended indicator over arrivals.
DEFAULT_TABLE = 1

TABLE_SHEET_NUMBERS = list(range(1, 34))  # sheets "1".."33" (all monthly time series)
COMUNA_SHEET_NAME = "34"                   # region/destino/comuna lookup, not a time series
INDEX_SHEET_NAME = "Índice"

TITLE_ROW = 3     # "Cuadro N.- <description>"
HEADER_ROW = 7    # "Región / Destino turístico", "Jul-16", "Ago-16", ...
DATA_START_ROW = 8

# Column A markers that signal "this row is a footnote, not data" -- checked
# case-insensitively against the start of the (stripped) cell text.
STOP_MARKERS = ("FUENTE", "NOTA", "/P", "/R", "CUADRO")

MONTH_ABBR_TO_NUM = {
    "Ene": 1, "Feb": 2, "Mar": 3, "Abr": 4, "May": 5, "Jun": 6,
    "Jul": 7, "Ago": 8, "Sept": 9, "Oct": 10, "Nov": 11, "Dic": 12,
}

# ---------------------------------------------------------------------------

RAW_DIR = Path(__file__).resolve().parent.parent.parent / "raw" / "chile_ine_tourism"
PROCESSED_DIR = Path(__file__).resolve().parent.parent.parent / "processed" / "americas"
CACHED_XLSX_PATH = RAW_DIR / "series-mensuales-de-julio-2016-a-la-fecha.xlsx"
OUTPUT_FILENAME = "chile_ine_tourism_monthly.csv"
COMUNA_OUTPUT_FILENAME = "chile_ine_destino_turistico_comunas.csv"

ATTRIBUTION = f"Instituto Nacional de Estadísticas (INE) Chile, EMAT -- {SOURCE_PAGE_URL}"


def _download_live() -> None:
    """GET SOURCE_XLSX_URL and overwrite CACHED_XLSX_PATH. A plain default
    User-Agent gets a 403 from ine.gob.cl (bot-blocking) -- a browser-like
    one works, same fix used for SimpleMaps elsewhere in this project."""
    headers = {
        "User-Agent": (
            "Mozilla/5.0 (Macintosh; Intel Mac OS X 10_15_7) "
            "AppleWebKit/537.36 (KHTML, like Gecko) Chrome/124.0 Safari/537.36"
        )
    }
    resp = requests.get(SOURCE_XLSX_URL, headers=headers, timeout=60)
    resp.raise_for_status()
    CACHED_XLSX_PATH.write_bytes(resp.content)


def download_workbook(force: bool = False) -> Path:
    """Fetch the EMAT workbook, preferring a live download but falling back
    to the cached copy in raw/ if the request fails (as it does in this
    sandbox -- ine.gob.cl is network-blocked here, see docstring). With
    force=True there's no fallback -- a failed request raises, since the
    whole point of --force-download is to bypass the cache."""
    RAW_DIR.mkdir(parents=True, exist_ok=True)

    if force:
        _download_live()
        print(f"Downloaded (forced) -> {CACHED_XLSX_PATH}")
        return CACHED_XLSX_PATH

    try:
        _download_live()
        print(f"Downloaded -> {CACHED_XLSX_PATH}")
        return CACHED_XLSX_PATH
    except requests.RequestException as e:
        if CACHED_XLSX_PATH.exists():
            print(f"Live download failed ({e}) -- using cached copy: {CACHED_XLSX_PATH}")
            return CACHED_XLSX_PATH
        raise


def load_workbook(path: Path) -> openpyxl.Workbook:
    """Load with data_only=True (resolve formulas to values) and NOT
    read_only -- read_only mode drops cell formatting, and this script
    needs `font.bold` to tell region rows from destino turístico rows."""
    return openpyxl.load_workbook(path, data_only=True, read_only=False)


def list_tables(wb: openpyxl.Workbook) -> dict[int, str]:
    """Parse the Índice sheet into {table_number: description}. Table 34
    is included (it's listed there too) even though it's handled
    separately from 1-33 elsewhere in this script."""
    ws = wb[INDEX_SHEET_NAME]
    tables = {}
    for row in ws.iter_rows(values_only=True):
        val = row[0]
        if not isinstance(val, str):
            continue
        m = re.match(r"^(\d+)\.-\s*(.+)$", val.strip())
        if m:
            tables[int(m.group(1))] = m.group(2).strip()
    return tables


def parse_month_columns(ws) -> list[tuple[int, str]]:
    """Return [(column_index, 'YYYY-MM'), ...] for every header cell in
    HEADER_ROW that looks like a Spanish month abbreviation + 2-digit year
    (e.g. 'Jul-16', 'Sept-25', 'Abr-26/R' -- the '/R' rectified-figures
    suffix is stripped). Trailing computed columns ('Variación (%) en
    doce meses', etc.) don't match the month-name lookup and are skipped."""
    month_cols = []
    for col_idx in range(2, ws.max_column + 1):
        label = ws.cell(row=HEADER_ROW, column=col_idx).value
        if not isinstance(label, str) or "-" not in label:
            continue
        mon_part, _, year_part = label.strip().partition("-")
        if mon_part not in MONTH_ABBR_TO_NUM:
            continue
        year_part = year_part.split("/")[0].strip()
        if not (year_part.isdigit() and len(year_part) == 2):
            continue
        ref_date = f"20{year_part}-{MONTH_ABBR_TO_NUM[mon_part]:02d}"
        month_cols.append((col_idx, ref_date))
    return month_cols


def _is_stop_row(label) -> bool:
    if not isinstance(label, str) or label.strip() == "":
        return True
    upper = label.strip().upper()
    return any(upper.startswith(marker) for marker in STOP_MARKERS)


def parse_table_sheet(ws, table_number: int, table_name: str) -> pd.DataFrame:
    """Parse one 'Cuadro' sheet (1-33) into tidy long format: one row per
    (region/destino turístico, month). `level` is 'national', 'region', or
    'destino' -- region rows and "Total nacional" are bold in the source,
    destino turístico rows nested under a region are not (see docstring)."""
    month_cols = parse_month_columns(ws)

    rows = []
    current_region = None
    r = DATA_START_ROW
    while True:
        label = ws.cell(row=r, column=1).value
        if _is_stop_row(label):
            break

        label = label.strip()
        is_bold = bool(ws.cell(row=r, column=1).font.bold)

        if label == "Total nacional":
            level, region, destino = "national", None, None
        elif is_bold:
            level, region, destino = "region", label, None
            current_region = label
        else:
            level, region, destino = "destino", current_region, label

        for col_idx, ref_date in month_cols:
            value = ws.cell(row=r, column=col_idx).value
            value = value if isinstance(value, (int, float)) else None
            rows.append(
                {
                    "table_number": table_number,
                    "table_name": table_name,
                    "level": level,
                    "region": region,
                    "destino_turistico": destino,
                    "ref_date": ref_date,
                    "value": value,
                }
            )
        r += 1

    return pd.DataFrame(rows)


def parse_comuna_mapping(wb: openpyxl.Workbook) -> pd.DataFrame:
    """Parse sheet 34 (region -> destino turístico -> comuna) into a tidy
    lookup table. Region/destino turístico cells are blank on continuation
    rows (no merged cells in the source), so this forward-fills them."""
    ws = wb[COMUNA_SHEET_NAME]
    rows = []
    region, destino = None, None
    r = HEADER_ROW  # data starts right after the header row on this sheet ("Región", "Destino turístico", "Comuna")
    while True:
        a = ws.cell(row=r, column=1).value
        b = ws.cell(row=r, column=2).value
        c = ws.cell(row=r, column=3).value
        if a is None and b is None and c is None:
            break
        region = a.strip() if isinstance(a, str) else region
        destino = b.strip() if isinstance(b, str) else destino
        comuna = c.strip() if isinstance(c, str) else c
        rows.append({"region": region, "destino_turistico": destino, "comuna": comuna})
        r += 1

    return pd.DataFrame(rows)


def build_datasets(wb: openpyxl.Workbook, tables: list[int]) -> tuple[pd.DataFrame, pd.DataFrame]:
    table_index = list_tables(wb)
    frames = []
    for t in tables:
        ws = wb[str(t)]
        name = table_index.get(t, f"Cuadro {t}")
        frames.append(parse_table_sheet(ws, t, name))

    long_df = pd.concat(frames, ignore_index=True)
    long_df = long_df.sort_values(["table_number", "ref_date", "level"]).reset_index(drop=True)

    comuna_df = parse_comuna_mapping(wb)
    return long_df, comuna_df


def write_outputs(long_df: pd.DataFrame, comuna_df: pd.DataFrame) -> tuple[Path, Path]:
    PROCESSED_DIR.mkdir(parents=True, exist_ok=True)
    long_path = PROCESSED_DIR / OUTPUT_FILENAME
    comuna_path = PROCESSED_DIR / COMUNA_OUTPUT_FILENAME
    long_df.to_csv(long_path, index=False)
    comuna_df.to_csv(comuna_path, index=False)
    return long_path, comuna_path


def fetch_dataset(tables: list[int], force_download: bool = False) -> tuple[Path, Path]:
    xlsx_path = download_workbook(force=force_download)
    print(f"Reading {xlsx_path} ...")
    wb = load_workbook(xlsx_path)

    long_df, comuna_df = build_datasets(wb, tables)
    long_path, comuna_path = write_outputs(long_df, comuna_df)

    print(f"Wrote {len(long_df)} rows across {len(tables)} table(s) -> {long_path}")
    print(f"Wrote {len(comuna_df)} region/destino/comuna rows -> {comuna_path}")
    return long_path, comuna_path


def main():
    parser = argparse.ArgumentParser(description=__doc__, formatter_class=argparse.RawDescriptionHelpFormatter)
    parser.add_argument(
        "--table", type=int, default=None,
        help=f"Single Cuadro number (1-33) to parse. Default: {DEFAULT_TABLE} "
        "(overnight stays, total). Ignored if --all-tables is set.",
    )
    parser.add_argument(
        "--all-tables", action="store_true",
        help="Parse every Cuadro 1-33 into one long CSV instead of just one.",
    )
    parser.add_argument(
        "--list-tables", action="store_true",
        help="Print all 34 table numbers/titles from the Índice sheet and exit -- no download/parse/write.",
    )
    parser.add_argument("--force-download", action="store_true", help="Bypass the cached raw/ copy.")
    args = parser.parse_args()

    if args.list_tables:
        xlsx_path = download_workbook(force=args.force_download)
        wb = load_workbook(xlsx_path)
        for num, desc in sorted(list_tables(wb).items()):
            print(f"{num:>2}. {desc}")
        return

    if args.table is not None and not (1 <= args.table <= 33):
        parser.error("--table must be between 1 and 33 (34 is the comuna lookup, handled automatically)")

    tables = TABLE_SHEET_NUMBERS if args.all_tables else [args.table or DEFAULT_TABLE]
    fetch_dataset(tables=tables, force_download=args.force_download)


if __name__ == "__main__":
    main()
